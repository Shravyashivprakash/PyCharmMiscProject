# g7_logic.py
import re
import html
import pandas as pd
from datetime import datetime
from docx import Document
import argparse
from openpyxl import load_workbook
import numpy as np
import pytest
import os

import config
from io_utils import append_results


__all__ = [
    "parse_requirements",
    "append_result",
    "if_else_syntax_check",
    "for_condition_syntax_check",
    "while_syntax_check",
    "switch_syntax_check",
    "edge_case_check",
    "checklist_presence",
    "div_by_zero_check",
    "infinite_loop_check",
    "null_pointer_check",
    "out_of_range_check",
    "var_input_analysis",
    "get_objective_for_check",
    "print_objective",
]

# -------------------------
# Objective mapping utilities
# -------------------------

# Central mapping: check name -> Objective ID
OBJECTIVE_MAP = {
    # Objective G 7.1
    "if_syntax_check": "Objective G 7.1",
    "for_condition_syntax_check": "Objective G 7.1",
    "while_syntax_check": "Objective G 7.1",
    "switch_syntax_check": "Objective G 7.1",

    # Objective G 7.2
    "edge_case_check": "Objective G 7.2",

    # Objective G 7.3
    "checklist_presence": "Objective G 7.3",

    # Objective G 7.7
    "div_by_zero_check": "Objective G 7.7",
    "infinite_loop_check": "Objective G 7.7",
    "null_pointer_check": "Objective G 7.7",
    "out_of_range_check": "Objective G 7.7",
    "var_input_analysis": "Objective G 7.7",

    # Not explicitly mapped in headings; leave None if not applicable
    # "var_input_analysis": None,
}


def get_objective_for_check(check: str) -> str | None:
    """
    Return the Objective ID string for a given check name, or None if no mapping exists.
    """
    return OBJECTIVE_MAP.get(check)


def print_objective(check: str) -> None:
    """
    Print 'Objective G 7.X' if the check belongs to Respective Objective G 7.X,
    else print 'No match found'. This satisfies the original requirement.
    """
    g71_checks = {
        "if_syntax_check",
        "for_condition_syntax_check",
        "while_syntax_check",
        "switch_syntax_check",
    }
    if check in g71_checks:
        print("Objective G 7.1")

    g72_checks = {
        "edge_case_check",
    }
    if check in g72_checks:
        print("Objective G 7.2")

    g73_checks = {
        "checklist_presence",
    }
    if check in g73_checks:
        print("Objective G 7.3")

    g77_checks = {
        "div_by_zero_check",
        "infinite_loop_check",
        "null_pointer_check",
        "out_of_range_check",
        "var_input_analysis",
    }
    if check in g77_checks:
        print("Objective G 7.7")


# -------------------------
# SRS reading utilities
# -------------------------

# ID    Description
# MRJ_SCU_STC_SRS_001   The software shall…
# [
#   ["ID", "Description"],
#   ["MRJ_SCU_STC_SRS_001", "The software shall…"]
# ]
def read_SRS_doc(path):
    doc = Document(path)
    rows = []
    for t in doc.tables:
        for r in t.rows:
            rows.append([c.text.strip() for c in r.cells])
    return rows


# Logic:
# Must contain "_"
# Must contain at least 1 number
# Must contain some letters
def looks_like_id(token):
    token = token.strip().strip(",.;:()[]")
    # If it has space(s), it's not an ID (it's likely a sentence)
    if " " in token or "\t" in token:
        return False
    return "_" in token and any(ch.isdigit() for ch in token) and any(ch.isalpha() for ch in token)


# Extract ID + Description from requirement document.
def parse_requirements(path):
    rows = read_SRS_doc(path)
    data = []  # ← store each row separately (allow duplicates)
    algo_data = []  # store only algorithm-related requirements

    current_id = None
    buffer = []

    for r in rows:
        if not r:
            continue

        first = r[0].strip()
        last = r[-1].strip()

        # Case 1: A NEW ID starts in this row
        if looks_like_id(first):
            # Save previous requirement (if any)
            if current_id:
                desc = " ".join(buffer).strip()
                data.append([current_id, desc])
                # check for "algorithm"
                if (
                        "algorithm" in desc.lower()
                        or "pseudocode" in desc.lower()
                        or "logic below" in desc.lower()
                        or "following logic" in desc.lower()
                ):
                    algo_data.append([current_id, desc])

            current_id = first
            buffer = [" ".join(r[1:]).strip()]  # start new description
            continue

        # Case 2: ID is at the end (rare but support it)
        elif looks_like_id(last):
            if current_id:
                desc = " ".join(buffer).strip()
                data.append([current_id, desc])
                if (
                        "algorithm" in desc.lower()
                        or "pseudocode" in desc.lower()
                        or "logic below" in desc.lower()
                        or "following logic" in desc.lower()
                ):
                    algo_data.append([current_id, desc])

            current_id = last
            buffer = [" ".join(r[:-1]).strip()]
            continue

        # Case 3: No ID → this row is continuation text
        if current_id:
            extra_text = " ".join(r).strip()
            if extra_text:
                buffer.append(extra_text)

    # Save last requirement
    if current_id:
        desc = " ".join(buffer).strip()
        data.append([current_id, desc])
        if (
                "algorithm" in desc.lower()
                or "pseudocode" in desc.lower()
                or "logic below" in desc.lower()
                or "following logic" in desc.lower()
        ):
            algo_data.append([current_id, desc])

    # Convert to DataFrame
    df = pd.DataFrame(data, columns=["ID", "DESC"])
    algo_df = pd.DataFrame(algo_data, columns=["ID", "DESC"])

    # Mark rows in df that are NOT in algo_df
    df["IS_NOT_ALGO"] = ~df.apply(tuple, axis=1).isin(algo_df.apply(tuple, axis=1))
    non_algo_df = df[df["IS_NOT_ALGO"]]

    # Save to Excel with multiple sheets
    with pd.ExcelWriter("requirements_output1.xlsx", engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="All_Requirements", index=False)
        algo_df.to_excel(writer, sheet_name="Algorithm_Requirements", index=False)
        non_algo_df.to_excel(writer, sheet_name="Non_Algorithm_Requirements", index=False)

    return algo_df


# -------------------------------------------------------------------
# Enhanced result appender: per-ID summary + detailed rows
# -------------------------------------------------------------------
def append_result(result):
    """
    Enhanced result appender that:
      - Extracts requirement IDs from 'details' strings (expects 'at ID [XYZ]').
      - Classifies severity per detail ('Error' -> Failed, 'Caution/Notice/Warning' -> Warning, else Info).
      - Computes aggregate status per ID (Failed > Warning > Passed).
      - Writes summary to 'Results' sheet and all messages to 'Results_Details' sheet.
      - Populates 'Objective ID' for each row based on the check name.

    Required fields in 'result':
      - result['check']: str
      - result['status']: str (overall status, used for IDs with no specific messages)
      - result['details']: list[str]
      - result['timestamp']: str
      - OPTIONAL result['ids']: list of all req IDs processed (to mark Passed if no messages)
    """
    required = ("check", "status", "details", "timestamp")
    if not all(k in result for k in required):
        raise ValueError("result must have keys: 'check', 'status', 'details', 'timestamp'")

    check_name = result["check"]
    overall = result["status"]
    details = result["details"] or []
    timestamp = result["timestamp"]
    all_ids = result.get("ids", [])  # list of all IDs for this check, used to mark Passed when no issues

    # Determine Objective ID and print if it is G 7.1 (per original requirement)
    objective_id = get_objective_for_check(check_name)
    print_objective(check_name)  # prints "Objective G 7.1" for G 7.1 checks; else "No match found"

    # Regex to extract ID from messages: expects "... at ID [XYZ] ..."
    id_rx = re.compile(r'at\s+ID\s*\[\s*(?P<ID>[^\]]+)\s*\]', re.IGNORECASE)

    def classify_severity(msg: str) -> str:
        m = msg.strip()
        if m.startswith("Error:"):
            return "Failed"
        if m.startswith("Caution:") or m.startswith("Notice:") or m.startswith("Warning:"):
            return "Warning"
        return "Info"

    # Collect detail rows per ID
    messages_by_id = {}  # {ID: [(msg, severity), ...]}
    global_msgs = []  # messages with no ID (e.g., "Scanned N requirements...")
    for msg in details:
        m = id_rx.search(msg)
        if m:
            req_id = m.group("ID").strip()
            severity = classify_severity(msg)
            messages_by_id.setdefault(req_id, []).append((msg, severity))
        else:
            global_msgs.append(msg)

    # Aggregate status per ID (Failed > Warning > Passed)
    def aggregate_status(detail_items):
        severities = {sev for (_, sev) in detail_items}
        if "Failed" in severities:
            return "Failed"
        if "Warning" in severities:
            return "Warning"
        return "Passed"  # Info-only messages → Passed

    summary_rows = []
    detail_rows = []

    # IDs with explicit messages
    for req_id, items in messages_by_id.items():
        agg = aggregate_status(items)
        summary_rows.append({
            "Objective ID": objective_id,
            "ID": req_id,
            "Check": check_name,
            "Status": agg,
            "Timestamp": timestamp
        })
        for msg, sev in items:
            detail_rows.append({
                "Objective ID": objective_id,
                "ID": req_id,
                "Check": check_name,
                "Status": sev,
                "Details": msg,
                "Timestamp": timestamp
            })

    # IDs with no messages: mark as Passed (no evidence against)
    default_status = "Passed"
    for req_id in all_ids:
        if req_id not in messages_by_id:
            summary_rows.append({
                "Objective ID": objective_id,
                "ID": req_id,
                "Check": check_name,
                "Status": default_status,
                "Timestamp": timestamp
            })

    # Add global messages to details (ID = '__GLOBAL__')
    for msg in global_msgs:
        detail_rows.append({
            "Objective ID": objective_id,
            "ID": "__GLOBAL__",
            "Check": check_name,
            "Status": classify_severity(msg),
            "Details": msg,
            "Timestamp": timestamp
        })

    # Build dataframes
    summary_df = pd.DataFrame(summary_rows, columns=["Objective ID", "ID", "Check", "Status", "Timestamp"])
    details_df = pd.DataFrame(detail_rows, columns=["Objective ID", "ID", "Check", "Status", "Details", "Timestamp"])

    excel_file = "Algorithm_analysis_result.xlsx"
    summary_sheet = "Results"
    details_sheet = "Results_Details"

    # Write/append to Excel
    if not os.path.exists(excel_file):
        with pd.ExcelWriter(excel_file, mode="w", engine="xlsxwriter") as writer:
            summary_df.to_excel(writer, sheet_name=summary_sheet, index=False, header=True)
            details_df.to_excel(writer, sheet_name=details_sheet, index=False, header=True)
        print(
            f"Created {excel_file} and Results appended to {excel_file} with sheets '{summary_sheet}'(summary) and '{details_sheet}'(details) for check 'if_else_syntax_check'")
        return

    # Append to existing sheets
    book = load_workbook(excel_file)

    def append_df(writer, sheet_name, df):
        if sheet_name in book.sheetnames:
            ws = book[sheet_name]
            startrow = ws.max_row  # append below last row
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
        else:
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

    with pd.ExcelWriter(excel_file, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
        append_df(writer, summary_sheet, summary_df)
        append_df(writer, details_sheet, details_df)

    print(
        f"Results appended to {excel_file} → '{summary_sheet}' (summary) and '{details_sheet}' (details) for check '{check_name}'")


## Objective G 7.1 ##

def if_else_syntax_check(algo_df):
    """
    Unified IF/ELSEIF/ELSE syntax checker for algo_df['DESC'] ...
    """
    results = {
        'check': "if_syntax_check",
        'status': "Passed",
        'details': [],
        'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'ids': list(pd.unique(algo_df['ID']))  # <-- Per-ID tracking
    }

    # Structural check
    total_algos = len(algo_df)
    if total_algos == 0:
        results['status'] = "Failed: algo_df is empty"
        append_result(results)
        ##return results

    # --- Token regex (case-insensitive) ---
    tok_re = re.compile(
        r'(?P<IF>\bIF\b)|'
        r'(?P<ELSEIF>\bELSE\s*IF\b|\bELSEIF\b)|'
        r'(?P<ELSE>\bELSE\b)|'
        r'(?P<THEN>\bTHEN\b)|'
        r'(?P<ENDIF>\bEND\s*IF\b|\bENDIF\b)',
        re.IGNORECASE
    )

    comparator_re = re.compile(r'(==|!=|>=|<=|>|<|&gt;=|&lt;=|&gt;|&lt;|&amp;gt;=|&amp;lt;=|&amp;gt;|&amp;lt;)')
    single_equals_re = re.compile(r'(?<![=!<>])=(?!=)')
    ambiguous_ops_re = re.compile(r'\b(EQ|NE|GT|LT|GE|LE)\b', re.IGNORECASE)

    def normalize_line(s: str) -> str:
        s = html.unescape(s)
        s = s.replace('≥', '&gt;=').replace('≤', '&lt;=').replace('≠', '!=')
        s = s.replace('&amp;eq;', '==').replace('&amp;amp;eq;', '==')
        return s

    def validate_condition(cond_text: str, req_id, line_num, context: str):
        text = cond_text.strip()
        if text == '':
            results['status'] = "Failed"
            results['details'].append(
                f"Error: Empty {context} condition at ID [{req_id}], Line {line_num}."
            )
            return
        # parentheses
        stack_paren = 0
        for ch in text:
            if ch == '(':
                stack_paren += 1
            elif ch == ')':
                stack_paren -= 1
                if stack_paren < 0:
                    results['status'] = "Failed"
                    results['details'].append(
                        f"Error: Unbalanced parentheses in {context} condition at ID [{req_id}], Line {line_num}: '{text}'"
                    )
                    break
        if stack_paren != 0:
            results['status'] = "Failed"
            results['details'].append(
                f"Error: Unbalanced parentheses in {context} condition at ID [{req_id}], Line {line_num}: '{text}'"
            )

        if single_equals_re.search(text):
            results['status'] = "Failed"
            results['details'].append(
                f"Error: Assignment '=' used in {context} condition (use '==') at ID [{req_id}], Line {line_num}: '{text}'"
            )

        if not comparator_re.search(text):
            results['status'] = "Failed"
            results['details'].append(
                f"Error: No comparison operator in {context} condition at ID [{req_id}], Line {line_num}: '{text}'"
            )

        if ambiguous_ops_re.search(text):
            results['details'].append(
                f"Notice: Textual comparator detected in {context} condition at ID [{req_id}], Line {line_num}: '{text}'. Prefer symbolic operators (==, !=, >, <, >=, <=)."
            )

    for index, row in algo_df.iterrows():
        desc_text = str(row.get('DESC', ''))
        req_id = row.get('ID', index)

        lines = desc_text.splitlines()
        stack = []

        for line_num, raw_line in enumerate(lines, 1):
            line = normalize_line(raw_line.rstrip())
            if not line.strip():
                if stack and stack[-1]['collecting'] in ('IF', 'ELSEIF'):
                    stack[-1]['cond_parts'].append('')
                continue

            tokens = list(tok_re.finditer(line))
            cursor = 0

            def add_body_segment(seg_text: str):
                text = seg_text.strip()
                if not text:
                    return
                if not stack:
                    return
                cb = stack[-1]['current_branch']
                if cb == 'IF':
                    stack[-1]['if_body_count'] += 1
                elif cb and cb.startswith('ELSEIF_'):
                    idx = int(cb.split('_')[1])
                    while len(stack[-1]['elseif_bodies']) <= idx:
                        stack[-1]['elseif_bodies'].append(0)
                    stack[-1]['elseif_bodies'][idx] += 1
                elif cb == 'ELSE':
                    stack[-1]['else_body_count'] += 1

            if not tokens:
                if stack and stack[-1]['collecting'] in ('IF', 'ELSEIF'):
                    stack[-1]['cond_parts'].append(line.strip())
                elif stack and stack[-1]['current_branch'] is not None:
                    add_body_segment(line[cursor:])
                continue

            for i, m in enumerate(tokens):
                kind = m.lastgroup
                start, end = m.start(), m.end()

                if stack and stack[-1]['collecting'] in ('IF', 'ELSEIF') and kind in ('IF', 'ELSEIF', 'ELSE', 'ENDIF'):
                    results['status'] = "Failed"
                    results['details'].append(
                        f"Error: {kind} encountered before THEN at ID [{req_id}], Line {line_num}: '{raw_line.strip()}'"
                    )
                    cursor = end
                    continue

                if kind == 'IF':
                    if stack and stack[-1]['current_branch'] is not None:
                        add_body_segment('nested_if')
                    stack.append({
                        'open_line': line_num,
                        'collecting': 'IF',
                        'cond_parts': [],
                        'else_seen': False,
                        'if_body_count': 0,
                        'elseif_bodies': [],
                        'else_body_count': 0,
                        'current_branch': None
                    })
                    cursor = end

                elif kind == 'ELSEIF':
                    if not stack:
                        results['status'] = "Failed"
                        results['details'].append(
                            f"Error: ELSEIF without an open IF at ID [{req_id}], Line {line_num}: '{raw_line.strip()}'"
                        )
                        cursor = end
                        continue
                    if stack[-1]['else_seen']:
                        results['status'] = "Failed"
                        results['details'].append(
                            f"Error: ELSEIF after ELSE is not allowed at ID [{req_id}], Line {line_num}: '{raw_line.strip()}'"
                        )
                        cursor = end
                        continue
                    prev = stack[-1]['current_branch']
                    if prev == 'IF' and stack[-1]['if_body_count'] == 0:
                        results['status'] = "Failed"
                        results['details'].append(
                            f"Error: Empty IF body before ELSEIF at ID [{req_id}], Line {line_num}: '{raw_line.strip()}'"
                        )
                    elif prev and prev.startswith('ELSEIF_'):
                        idx_prev = int(prev.split('_')[1])
                        if stack[-1]['elseif_bodies'][idx_prev] == 0:
                            results['status'] = "Failed"
                            results['details'].append(
                                f"Error: Empty ELSEIF body before another ELSEIF at ID [{req_id}], Line {line_num}: '{raw_line.strip()}'"
                            )
                    stack[-1]['collecting'] = 'ELSEIF'
                    stack[-1]['cond_parts'] = []
                    new_idx = len(stack[-1]['elseif_bodies'])
                    stack[-1]['current_branch'] = None
                    cursor = end

                elif kind == 'THEN':
                    if not stack or stack[-1]['collecting'] not in ('IF', 'ELSEIF'):
                        add_body_segment(line[cursor:start])
                        cursor = end
                        continue
                    stack[-1]['cond_parts'].append(line[cursor:start].strip())
                    cond_text = ' '.join(p for p in stack[-1]['cond_parts'] if p is not None).strip()
                    context = stack[-1]['collecting']
                    validate_condition(cond_text, req_id, line_num, context=context)
                    if context == 'IF':
                        stack[-1]['current_branch'] = 'IF'
                    else:
                        new_idx = len(stack[-1]['elseif_bodies'])
                        stack[-1]['elseif_bodies'].append(0)
                        stack[-1]['current_branch'] = f'ELSEIF_{new_idx}'
                    stack[-1]['collecting'] = None
                    stack[-1]['cond_parts'] = []
                    cursor = end

                elif kind == 'ELSE':
                    if not stack:
                        results['status'] = "Failed"
                        results['details'].append(
                            f"Error: ELSE without an open IF at ID [{req_id}], Line {line_num}: '{raw_line.strip()}'"
                        )
                        cursor = end
                        continue
                    add_body_segment(line[cursor:start])
                    if stack[-1]['current_branch'] is None:
                        results['status'] = "Failed"
                        results['details'].append(
                            f"Error: ELSE encountered before THEN at ID [{req_id}], Line {line_num}: '{raw_line.strip()}'"
                        )
                        cursor = end
                        continue
                    if stack[-1]['else_seen']:
                        results['status'] = "Failed"
                        results['details'].append(
                            f"Error: Multiple ELSE branches are not allowed at ID [{req_id}], Line {line_num}: '{raw_line.strip()}'"
                        )
                        cursor = end
                        continue
                    stack[-1]['else_seen'] = True
                    stack[-1]['current_branch'] = 'ELSE'
                    cursor = end

                elif kind == 'ENDIF':
                    if not stack:
                        results['status'] = "Failed"
                        results['details'].append(
                            f"Error: END IF/ENDIF without a matching IF at ID [{req_id}], Line {line_num}: '{raw_line.strip()}'"
                        )
                        cursor = end
                        continue
                    add_body_segment(line[cursor:start])
                    if stack[-1]['current_branch'] is None:
                        results['status'] = "Failed"
                        results['details'].append(
                            f"Error: END IF encountered before a THEN at ID [{req_id}], opened Line {stack[-1]['open_line']}, closed Line {line_num}."
                        )
                    else:
                        if stack[-1]['if_body_count'] <= 0:
                            results['status'] = "Failed"
                            results['details'].append(
                                f"Error: Empty IF body before END IF at ID [{req_id}], opened Line {stack[-1]['open_line']}, closed Line {line_num}."
                            )
                        for i, cnt in enumerate(stack[-1]['elseif_bodies']):
                            if cnt <= 0:
                                results['status'] = "Failed"
                                results['details'].append(
                                    f"Error: Empty ELSEIF body #{i + 1} before END IF at ID [{req_id}], opened Line {stack[-1]['open_line']}, closed Line {line_num}."
                                )
                        if stack[-1]['else_seen'] and stack[-1]['else_body_count'] <= 0:
                            results['status'] = "Failed"
                            results['details'].append(
                                f"Error: Empty ELSE body before END IF at ID [{req_id}], opened Line {stack[-1]['open_line']}, closed Line {line_num}."
                            )
                    stack.pop()
                    cursor = end

            if stack and stack[-1]['collecting'] is None and cursor < len(line):
                add_body_segment(line[cursor:])
            if stack and stack[-1]['collecting'] in ('IF', 'ELSEIF') and cursor < len(line):
                stack[-1]['cond_parts'].append(line[cursor:].strip())

        if stack:
            open_lines = ", ".join(str(b['open_line']) for b in stack)
            results['status'] = "Failed"
            results['details'].append(
                f"Error: Unclosed IF block(s) at ID [{req_id}]; opened at line(s): {open_lines}"
            )

    if not results['details']:
        results['details'].append(f"Scanned {total_algos} requirements; IF/ELSEIF/ELSE syntax looks good.")
    append_result(results)
    ##return results


def for_condition_syntax_check(algo_df, strict_uppercase=True, lookahead_limit=3):
    """
    Validate ONLY the FOR header condition in algo_df['DESC'] ...
    """
    results = {
        'check': "for_condition_syntax_check",
        'status': "Passed",
        'details': [],
        'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'ids': list(pd.unique(algo_df['ID']))  # <-- Per-ID tracking
    }

    total_algos = len(algo_df)
    if total_algos == 0:
        results['status'] = "Failed: algo_df is empty"
        append_result(results)
        ##return results

    tok_re = re.compile(r'(?P<FOR>\bFOR\b)|(?P<DO>\bDO\b)', re.IGNORECASE)
    comparator_re = re.compile(r'(==|!=|>=|<=|>|<)')
    single_equals_re = re.compile(r'(?<![=!<>])=(?!=)')
    for_range_re = re.compile(r'\b(TO|DOWNTO|IN)\b', re.IGNORECASE)
    ambiguous_ops_re = re.compile(r'\b(EQ|NE|GT|LT|GE|LE)\b', re.IGNORECASE)

    def normalize_line(s: str) -> str:
        for _ in range(2):
            s = html.unescape(s)
        s = s.replace('≥', '&gt;=').replace('≤', '&lt;=').replace('≠', '!=')
        s = s.replace('&amp;eq;', '==').replace('&amp;amp;eq;', '==')
        return s

    def validate_parens(text: str, req_id, line_num, context: str):
        stack = 0
        for ch in text:
            if ch == '(':
                stack += 1
            elif ch == ')':
                stack -= 1
            if stack < 0:
                results['status'] = "Failed"
                results['details'].append(
                    f"Error: Unbalanced parentheses in {context} at ID [{req_id}], Line {line_num}: '{text.strip()}'"
                )
                return
        if stack != 0:
            results['status'] = "Failed"
            results['details'].append(
                f"Error: Unbalanced parentheses in {context} at ID [{req_id}], Line {line_num}: '{text.strip()}'"
            )

    def validate_for_condition(cond_text: str, req_id, line_num):
        text = cond_text.strip()
        if text == '':
            results['status'] = "Failed"
            results['details'].append(
                f"Error: Empty FOR condition at ID [{req_id}], Line {line_num}."
            )
            return

        validate_parens(text, req_id, line_num, context="FOR condition")

        has_comp = bool(comparator_re.search(text))
        has_range = bool(for_range_re.search(text))

        if single_equals_re.search(text) and not has_range:
            results['status'] = "Failed"
            results['details'].append(
                f"Error: Assignment '=' used in FOR condition without a range keyword (use '==' for comparison) at ID [{req_id}], Line {line_num}: '{text}'"
            )

        if not (has_comp or has_range):
            results['status'] = "Failed"
            results['details'].append(
                f"Error: No comparator or range keyword in FOR condition at ID [{req_id}], Line {line_num}: '{text}'. Expected one of (==, !=, >=, <=, >, <) or TO/DOWNTO/IN."
            )

        if ambiguous_ops_re.search(text):
            results['details'].append(
                f"Notice: Textual comparator detected in FOR condition at ID [{req_id}], Line {line_num}: '{text}'. Prefer symbolic operators (==, !=, >, <, >=, <=)."
            )

    def is_probable_for_header(lines_norm, start_idx, after_for_text):
        seg = after_for_text.strip()
        if comparator_re.search(seg) or for_range_re.search(seg):
            return True
        if strict_uppercase:
            if 'DO' in seg.split():
                return True
        else:
            if re.search(r'\bDO\b', seg, re.IGNORECASE):
                return True
        for k in range(1, min(lookahead_limit + 1, len(lines_norm) - start_idx)):
            nxt = lines_norm[start_idx + k].strip()
            if comparator_re.search(nxt) or for_range_re.search(nxt):
                return True
            if strict_uppercase:
                if 'DO' in nxt.split():
                    return True
            else:
                if re.search(r'\bDO\b', nxt, re.IGNORECASE):
                    return True
        return False

    for index, row in algo_df.iterrows():
        desc_text = str(row.get('DESC', ''))
        req_id = row.get('ID', index)

        raw_lines = desc_text.splitlines()
        lines = [normalize_line(ln.rstrip()) for ln in raw_lines]

        for line_num, line in enumerate(lines, 1):
            tokens = list(tok_re.finditer(line))
            if not tokens:
                continue

            for m in tokens:
                kind = m.lastgroup
                start, end = m.start(), m.end()
                matched_text = m.group()

                if kind == 'FOR':
                    if strict_uppercase and matched_text != 'FOR':
                        continue
                    after_for = line[end:]
                    if not is_probable_for_header(lines, line_num - 1, after_for):
                        continue
                    if strict_uppercase:
                        do_match_same = re.search(r'\bDO\b', after_for)
                    else:
                        do_match_same = re.search(r'\bDO\b', after_for, re.IGNORECASE)

                    if do_match_same:
                        cond_text = after_for[:do_match_same.start()].strip()
                        validate_for_condition(cond_text, req_id, line_num)
                        continue

                    cond_parts = []
                    if after_for.strip():
                        cond_parts.append(after_for.strip())

                    found_do = False
                    for ahead_idx in range(line_num, min(line_num + lookahead_limit, len(lines))):
                        ahead_line = lines[ahead_idx]
                        if strict_uppercase:
                            m_do = re.search(r'\bDO\b', ahead_line)
                        else:
                            m_do = re.search(r'\bDO\b', ahead_line, re.IGNORECASE)

                        if m_do:
                            prefix = ahead_line[:m_do.start()].strip()
                            if prefix:
                                cond_parts.append(prefix)
                            cond_text = ' '.join(part for part in cond_parts).strip()
                            validate_for_condition(cond_text, req_id, ahead_idx + 0)
                            found_do = True
                            break
                        else:
                            if ahead_line.strip():
                                cond_parts.append(ahead_line.strip())

                    if not found_do:
                        continue

    if not results['details']:
        results['details'].append(
            f"Scanned {total_algos} requirements; FOR conditions look good (strict_uppercase={strict_uppercase}; ignored non-control uses of 'for')."
        )
    append_result(results)
    ##return results


def while_syntax_check(algo_df):
    """
    Unified WHILE syntax checker ...
    """
    results = {
        'check': "while_syntax_check",
        'status': "Passed",
        'details': [],
        'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'ids': list(pd.unique(algo_df['ID']))  # <-- Per-ID tracking
    }

    total_algos = len(algo_df)
    if total_algos == 0:
        results['status'] = "Failed: algo_df is empty"
        append_result(results)
        ##return results

    tok_re = re.compile(
        r'(?P<WHILE>\bWHILE\b)|'
        r'(?P<DO>\bDO\b)|'
        r'(?P<END_WHILE>\bEND\s*WHILE\b|\bENDWHILE\b)|'
        r'(?P<NEST_IF>\bIF\b)|(?P<NEST_FOR>\bFOR\b)|(?P<NEST_WHILE2>\bWHILE\b)|(?P<NEST_SWITCH>\bSWITCH\b)',
        re.IGNORECASE
    )

    comparator_re = re.compile(r'(==|!=|>=|<=|>|<)')
    single_equals_re = re.compile(r'(?<![=!<>])=(?!=)')
    ambiguous_ops_re = re.compile(r'\b(EQ|NE|GT|LT|GE|LE)\b', re.IGNORECASE)
    boolean_literal_re = re.compile(r'^\s*\(?\s*(TRUE|FALSE)\s*\)?\s*$', re.IGNORECASE)

    def normalize_line(s: str) -> str:
        for _ in range(2):
            s = html.unescape(s)
        s = s.replace('≥', '&gt;=').replace('≤', '&lt;=').replace('≠', '!=')
        s = s.replace('&amp;eq;', '==').replace('&amp;amp;eq;', '==')
        return s

    def validate_parens(text: str, req_id, line_num, context: str):
        stack = 0
        for ch in text:
            if ch == '(':
                stack += 1
            elif ch == ')':
                stack -= 1
            if stack < 0:
                results['status'] = "Failed"
                results['details'].append(
                    f"Error: Unbalanced parentheses in {context} at ID [{req_id}], Line {line_num}: '{text.strip()}'"
                )
                return
        if stack != 0:
            results['status'] = "Failed"
            results['details'].append(
                f"Error: Unbalanced parentheses in {context} at ID [{req_id}], Line {line_num}: '{text.strip()}'"
            )

    def validate_while_condition(cond_text: str, req_id, line_num):
        text = cond_text.strip()
        if text == '':
            results['status'] = "Failed"
            results['details'].append(
                f"Error: Empty WHILE condition at ID [{req_id}], Line {line_num}."
            )
            return

        validate_parens(text, req_id, line_num, context="WHILE condition")

        if single_equals_re.search(text):
            results['status'] = "Failed"
            results['details'].append(
                f"Error: Assignment '=' used in WHILE condition (use '==' for equality) at ID [{req_id}], Line {line_num}: '{text}'"
            )

        has_comp = bool(comparator_re.search(text))
        is_boolean_literal = bool(boolean_literal_re.match(text))

        if not (has_comp or is_boolean_literal):
            results['status'] = "Failed"
            results['details'].append(
                f"Error: WHILE condition must be a comparator expression or a boolean literal (TRUE/FALSE) at ID [{req_id}], Line {line_num}: '{text}'"
            )

        if ambiguous_ops_re.search(text) and not is_boolean_literal:
            results['details'].append(
                f"Notice: Textual comparator detected in WHILE condition at ID [{req_id}], Line {line_num}: '{text}'. Prefer symbolic operators (==, !=, >, <, >=, <=)."
            )

    for index, row in algo_df.iterrows():
        desc_text = str(row.get('DESC', ''))
        req_id = row.get('ID', index)

        lines = desc_text.splitlines()
        stack = []

        for line_num, raw_line in enumerate(lines, 1):
            line = normalize_line(raw_line.rstrip())
            if not line.strip():
                if stack and stack[-1].get('collecting') == 'COND':
                    stack[-1]['cond_parts'].append('')
                continue

            tokens = list(tok_re.finditer(line))
            cursor = 0

            def add_body_segment(seg_text: str):
                text = seg_text.strip()
                if not text or not stack:
                    return
                top = stack[-1]
                if top['type'] == 'WHILE' and top.get('collecting') is None:
                    top['body_count'] += 1

            def finalize_segment_until(start):
                if stack and stack[-1].get('collecting') is None and start > cursor:
                    add_body_segment(line[cursor:start])

            if not tokens:
                if stack:
                    top = stack[-1]
                    if top.get('collecting') == 'COND':
                        top['cond_parts'].append(line.strip())
                    else:
                        add_body_segment(line[cursor:])
                continue

            for m in tokens:
                kind = m.lastgroup
                start, end = m.start(), m.end()

                if stack and stack[-1].get('collecting') == 'COND':
                    if kind not in ('DO',):
                        results['status'] = "Failed"
                        results['details'].append(
                            f"Error: {kind} encountered before DO in WHILE header at ID [{req_id}], Line {line_num}: '{raw_line.strip()}'"
                        )
                        cursor = end
                        continue

                if kind == 'WHILE':
                    if stack and stack[-1].get('collecting') is None:
                        finalize_segment_until(start)
                        add_body_segment('nested_WHILE')

                    stack.append({
                        'type': 'WHILE',
                        'open_line': line_num,
                        'collecting': 'COND',
                        'cond_parts': [],
                        'body_count': 0
                    })
                    cursor = end

                elif kind == 'DO':
                    if not stack or stack[-1]['type'] != 'WHILE' or stack[-1].get('collecting') != 'COND':
                        finalize_segment_until(start)
                        cursor = end
                        continue

                    stack[-1]['cond_parts'].append(line[cursor:start].strip())
                    cond_text = ' '.join(p for p in stack[-1]['cond_parts'] if p is not None).strip()
                    validate_while_condition(cond_text, req_id, line_num)
                    stack[-1]['collecting'] = None
                    stack[-1]['cond_parts'] = []
                    cursor = end

                elif kind == 'END_WHILE':
                    if not stack or stack[-1]['type'] != 'WHILE':
                        results['status'] = "Failed"
                        results['details'].append(
                            f"Error: END WHILE/ENDWHILE without matching WHILE at ID [{req_id}], Line {line_num}: '{raw_line.strip()}'"
                        )
                        cursor = end
                        continue

                    finalize_segment_until(start)
                    if stack[-1]['body_count'] <= 0:
                        results['status'] = "Failed"
                        results['details'].append(
                            f"Error: Empty WHILE body before END WHILE at ID [{req_id}], opened Line {stack[-1]['open_line']}, closed Line {line_num}."
                        )
                    stack.pop()
                    cursor = end

                elif kind in ('NEST_IF', 'NEST_FOR', 'NEST_SWITCH', 'NEST_WHILE2'):
                    finalize_segment_until(start)
                    add_body_segment(f"nested_{kind}")
                    cursor = end

            if stack and stack[-1].get('collecting') is None and cursor < len(line):
                add_body_segment(line[cursor:])
            if stack and stack[-1].get('collecting') == 'COND' and cursor < len(line):
                stack[-1]['cond_parts'].append(line[cursor:].strip())

        if stack:
            open_desc = ", ".join(f"{b['type']}@{b['open_line']}" for b in stack)
            results['status'] = "Failed"
            results['details'].append(
                f"Error: Unclosed WHILE block(s) at ID [{req_id}]; opened: {open_desc}"
            )

    if not results['details']:
        results['details'].append(f"Scanned {total_algos} requirements; WHILE syntax looks good.")
    append_result(results)
    ##return results


def switch_syntax_check(algo_df):
    """
    Unified SWITCH syntax checker ...
    """
    results = {
        'check': "switch_syntax_check",
        'status': "Passed",
        'details': [],
        'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'ids': list(pd.unique(algo_df['ID']))  # <-- Per-ID tracking
    }

    total_algos = len(algo_df)
    if total_algos == 0:
        results['status'] = "Failed: algo_df is empty"
        append_result(results)
        ##return results

    tok_re = re.compile(
        r'(?P<SWITCH>\bSWITCH\b)|'
        r'(?P<CASE>\bCASE\b)|'
        r'(?P<DEFAULT>\bDEFAULT\b)|'
        r'(?P<THEN>\bTHEN\b)|'
        r'(?P<COLON>:)|'
        r'(?P<END_SWITCH>\bEND\s*SWITCH\b|\bENDSWITCH\b)|'
        r'(?P<NEST_IF>\bIF\b)|(?P<NEST_FOR>\bFOR\b)|(?P<NEST_WHILE>\bWHILE\b)|(?P<NEST_SWITCH2>\bSWITCH\b)',
        re.IGNORECASE
    )

    single_equals_re = re.compile(r'(?<![=!<>])=(?!=)')

    def normalize_line(s: str) -> str:
        for _ in range(2):
            s = html.unescape(s)
        s = s.replace('≥', '&gt;=').replace('≤', '&lt;=').replace('≠', '!=')
        s = s.replace('&amp;eq;', '==').replace('&amp;amp;eq;', '==')
        return s

    def validate_parens(text: str, req_id, line_num, context: str):
        stack = 0
        for ch in text:
            if ch == '(':
                stack += 1
            elif ch == ')':
                stack -= 1
            if stack < 0:
                results['status'] = "Failed"
                results['details'].append(
                    f"Error: Unbalanced parentheses in {context} at ID [{req_id}], Line {line_num}: '{text.strip()}'"
                )
                return
        if stack != 0:
            results['status'] = "Failed"
            results['details'].append(
                f"Error: Unbalanced parentheses in {context} at ID [{req_id}], Line {line_num}: '{text.strip()}'"
            )

    def validate_switch_expr(expr_text: str, req_id, line_num):
        t = expr_text.strip()
        if t == '':
            results['status'] = "Failed"
            results['details'].append(
                f"Error: Empty SWITCH expression at ID [{req_id}], Line {line_num}."
            )
            return
        validate_parens(t, req_id, line_num, context="SWITCH expression")
        if single_equals_re.search(t):
            results['status'] = "Failed"
            results['details'].append(
                f"Error: Assignment '=' used in SWITCH expression (use '==' if equality intended) at ID [{req_id}], Line {line_num}: '{t}'"
            )

    def validate_case_selector(sel_text: str, req_id, line_num):
        t = sel_text.strip()
        if t == '':
            results['status'] = "Failed"
            results['details'].append(
                f"Error: Empty CASE selector at ID [{req_id}], Line {line_num}."
            )
            return
        validate_parens(t, req_id, line_num, context="CASE selector")
        if single_equals_re.search(t):
            results['status'] = "Failed"
            results['details'].append(
                f"Error: Assignment '=' used in CASE selector (use '==' if equality intended) at ID [{req_id}], Line {line_num}: '{t}'"
            )

    for index, row in algo_df.iterrows():
        req_id = row.get('ID', index)
        desc_text = str(row.get('DESC', ''))

        lines = desc_text.splitlines()
        stack = []

        for line_num, raw_line in enumerate(lines, 1):
            line = normalize_line(raw_line.rstrip())
            if not line.strip():
                continue

            tokens = list(tok_re.finditer(line))
            cursor = 0

            def add_body_segment(seg_text: str):
                text = seg_text.strip()
                if not text or not stack:
                    return
                top = stack[-1]
                if top['type'] == 'SWITCH' and top.get('collecting') is None:
                    if top.get('current_branch', '').startswith('CASE_'):
                        idx = int(top['current_branch'].split('_')[1])
                        top['cases'][idx]['body_count'] += 1
                    elif top.get('current_branch') == 'DEFAULT':
                        top['default_body_count'] += 1

            def finalize_segment_until(start):
                if stack and stack[-1].get('collecting') is None and start > cursor:
                    add_body_segment(line[cursor:start])

            if not tokens:
                continue

            for m in tokens:
                kind = m.lastgroup
                start, end = m.start(), m.end()

                if stack and stack[-1].get('collecting') == 'SWITCH_EXPR':
                    if kind == 'CASE':
                        stack[-1]['cond_parts'].append(line[cursor:start].strip())
                        expr = ' '.join(p for p in stack[-1]['cond_parts']).strip()
                        validate_switch_expr(expr, req_id, line_num)
                        stack[-1]['collecting'] = None
                        stack[-1]['cond_parts'] = []
                        cursor = end
                    else:
                        results['status'] = "Failed"
                        results['details'].append(
                            f"Error: {kind} encountered before CASE after SWITCH at ID [{req_id}], Line {line_num}: '{raw_line.strip()}'"
                        )
                        cursor = end
                        continue

                if stack and stack[-1].get('collecting') == 'CASE':
                    if kind in ('THEN', 'COLON'):
                        stack[-1]['cond_parts'].append(line[cursor:start].strip())
                        selector = ' '.join(p for p in stack[-1]['cond_parts']).strip()
                        validate_case_selector(selector, req_id, line_num)
                        stack[-1]['collecting'] = None
                        stack[-1]['cond_parts'] = []
                        if stack[-1]['cases']:
                            if stack[-1]['cases'][-1]['cond'] is None:
                                stack[-1]['cases'][-1]['cond'] = selector
                        idx = len(stack[-1]['cases']) - 1
                        stack[-1]['current_branch'] = f'CASE_{idx}'
                        cursor = end
                        continue
                    elif kind in ('SWITCH', 'CASE', 'DEFAULT', 'END_SWITCH'):
                        results['status'] = "Failed"
                        results['details'].append(
                            f"Error: {kind} encountered before THEN/COLON after CASE at ID [{req_id}], Line {line_num}: '{raw_line.strip()}'"
                        )
                        cursor = end
                        continue

                if kind == 'SWITCH':
                    if stack and stack[-1].get('collecting') is None:
                        finalize_segment_until(start)
                        add_body_segment('nested_SWITCH')

                    stack.append({
                        'type': 'SWITCH',
                        'open_line': line_num,
                        'collecting': 'SWITCH_EXPR',
                        'cond_parts': [],
                        'cases': [],
                        'default_seen': False,
                        'default_body_count': 0,
                        'current_branch': None
                    })
                    cursor = end

                elif kind == 'CASE':
                    if not stack or stack[-1]['type'] != 'SWITCH':
                        results['status'] = "Failed"
                        results['details'].append(
                            f"Error: CASE without matching SWITCH at ID [{req_id}], Line {line_num}: '{raw_line.strip()}'"
                        )
                        cursor = end
                        continue
                    if stack[-1]['default_seen']:
                        results['status'] = "Failed"
                        results['details'].append(
                            f"Error: CASE after DEFAULT is not allowed at ID [{req_id}], Line {line_num}: '{raw_line.strip()}'"
                        )
                        cursor = end
                        continue
                    finalize_segment_until(start)
                    stack[-1]['collecting'] = 'CASE'
                    stack[-1]['cond_parts'] = []
                    stack[-1]['cases'].append({'cond': None, 'body_count': 0, 'start_line': line_num})
                    stack[-1]['current_branch'] = None
                    cursor = end

                elif kind in ('THEN', 'COLON'):
                    finalize_segment_until(start)
                    cursor = end
                    continue

                elif kind == 'DEFAULT':
                    if not stack or stack[-1]['type'] != 'SWITCH':
                        results['status'] = "Failed"
                        results['details'].append(
                            f"Error: DEFAULT without matching SWITCH at ID [{req_id}], Line {line_num}: '{raw_line.strip()}'"
                        )
                        cursor = end
                        continue
                    if stack[-1]['default_seen']:
                        results['status'] = "Failed"
                        results['details'].append(
                            f"Error: Multiple DEFAULT branches are not allowed at ID [{req_id}], Line {line_num}: '{raw_line.strip()}'"
                        )
                        cursor = end
                        continue
                    finalize_segment_until(start)
                    stack[-1]['default_seen'] = True
                    stack[-1]['current_branch'] = 'DEFAULT'
                    cursor = end

                elif kind == 'END_SWITCH':
                    if not stack or stack[-1]['type'] != 'SWITCH':
                        results['status'] = "Failed"
                        results['details'].append(
                            f"Error: END SWITCH/ENDSWITCH without matching SWITCH at ID [{req_id}], Line {line_num}: '{raw_line.strip()}'"
                        )
                        cursor = end
                        continue
                    finalize_segment_until(start)
                    top = stack[-1]
                    if len(top['cases']) == 0:
                        results['status'] = "Failed"
                        results['details'].append(
                            f"Error: SWITCH has no CASE branches at ID [{req_id}], opened Line {top['open_line']}, closed Line {line_num}."
                        )
                    for i, c in enumerate(top['cases'], 1):
                        if c['cond'] is None:
                            results['status'] = "Failed"
                            results['details'].append(
                                f"Error: CASE #{i} lacks selector/THEN at ID [{req_id}], opened Line {c['start_line']}, closed Line {line_num}."
                            )
                        if c['body_count'] <= 0:
                            results['status'] = "Failed"
                            results['details'].append(
                                f"Error: Empty CASE #{i} body before END SWITCH at ID [{req_id}], opened Line {c['start_line']}, closed Line {line_num}."
                            )
                    if top['default_seen'] and top['default_body_count'] <= 0:
                        results['status'] = "Failed"
                        results['details'].append(
                            f"Error: Empty DEFAULT body before END SWITCH at ID [{req_id}], opened Line {top['open_line']}, closed Line {line_num}."
                        )
                    stack.pop()
                    cursor = end

                elif kind in ('NEST_IF', 'NEST_FOR', 'NEST_WHILE', 'NEST_SWITCH2'):
                    finalize_segment_until(start)
                    add_body_segment(f"nested_{kind}")
                    cursor = end

            if stack and stack[-1].get('collecting') is None and cursor < len(line):
                add_body_segment(line[cursor:])
            if stack and stack[-1].get('collecting') in ('SWITCH_EXPR', 'CASE') and cursor < len(line):
                stack[-1]['cond_parts'].append(line[cursor:].strip())

        if stack:
            open_desc = ", ".join(f"{b['type']}@{b['open_line']}" for b in stack)
            results['status'] = "Failed"
            results['details'].append(
                f"Error: Unclosed SWITCH block(s) at ID [{req_id}]; opened: {open_desc}"
            )

    if not results['details']:
        results['details'].append(f"Scanned {total_algos} requirements; SWITCH syntax looks good.")
    append_result(results)
    ##return results


## Objective G 7.2 ##

def edge_case_check(algo_df):
    """
    Suite to test for DO-178 robustness term mentions in algorithm descriptions.
    """
    results = {
        'check': "edge_case_check",
        'status': "Passed",
        'details': [],
        'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'ids': list(pd.unique(algo_df['ID']))  # <-- Per-ID tracking
    }

    total_algos = len(algo_df)
    if total_algos == 0:
        results['status'] = "Failed: algo_df is empty"
        append_result(results)
        ##return results

    cats = {
        "Boundary/Range": ["min", "max", "range", "limit", "boundary", "upper", "lower", "overflow", "underflow",
                           "saturation", "clamp"],
        "Invalid Input": ["invalid", "null", "empty", "out of range", "negative", "zero", "format", "type", "nan"],
        "Timing/Timeout": ["timeout", "deadline", "latency", "period", "rate", "watchdog", "tick", "delay"],
        "Failure/Exceptions": ["error", "fault", "failure", "exception", "abort", "reset", "degraded", "safe state",
                               "recovery", "retry", "report"],
        "Resource/Init": ["memory", "stack", "leak", "resource", "alloc", "allocation", "initialization", "startup",
                          "power", "brown-out", "loss"],
        "Concurrency": ["race", "deadlock", "mutex", "lock", "priority inversion", "concurrent", "atomic"],
        "Interface/Protocol": ["message", "crc", "checksum", "protocol", "handshake", "header", "sequence",
                               "status code"],
        "Numerical": ["precision", "divide by zero", "rounding", "overflow", "underflow", "saturation"],
        "Mode/State": ["mode", "state", "transition", "unexpected", "inhibit", "failsafe", "fallback"]
    }

    cat_rx = {
        cat: [re.compile(r'\b' + re.escape(term) + r'\b', re.IGNORECASE) for term in terms]
        for cat, terms in cats.items()
    }

    for index, row in algo_df.iterrows():
        desc_text = str(row.get('DESC', '') or '')
        req_id = row.get('ID', f"Row{index}")

        if not desc_text.strip():
            results['details'].append(f"Notice: Empty DESC at ID [{req_id}]")
            continue

        txt = ' ' + desc_text.lower() + ' '
        matched = {}
        for cat, patterns in cat_rx.items():
            terms_hit = []
            for rx in patterns:
                if rx.search(txt):
                    term_literal = rx.pattern.replace(r'\b', '')
                    terms_hit.append(term_literal)
            if terms_hit:
                matched[cat] = sorted(set(terms_hit))

        if matched:
            if results['status'] == "Passed":
                results['status'] = "Failed"
            cat_list = ';'.join(matched.keys())
            term_detail = '; '.join([f"{k}: {', '.join(v)}" for k, v in matched.items()])
            results['details'].append(
                f"Notice: Edge-case categories hit at ID [{req_id}] → [{cat_list}] | terms: {term_detail}"
            )

    if results['status'] == "Passed":
        results['details'].append(
            f"Scanned {total_algos} requirements; no DO-178 edge-case terms found."
        )

    append_result(results)
    ##return results


# -------------------------
# Objective G 7.3 Coverage & checklist
# -------------------------

def checklist_presence(algo_df):
    """
    Suite to test for DO-178 robustness term mentions in algorithm descriptions.
    """
    results = {
        'check': "checklist_presence",
        'status': "Passed",
        'details': [],
        'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'ids': list(pd.unique(algo_df['ID']))  # <-- Per-ID tracking
    }

    total_algos = len(algo_df)
    if total_algos == 0:
        results['status'] = "Failed: algo_df is empty"
        append_result(results)
        ##return results

    checks = {
        "Numeric ranges/limits specified": ["range", "limit", "bounds", "min", "max", "boundary"],
        "Precision/representation defined": ["precision", "floating point", "fixed point", "q-format", "rounding",
                                             "truncate"],
        "Overflow/underflow/saturation handling": ["overflow", "underflow", "saturation", "clip", "clipping",
                                                   "wraparound"],
        "Error bounds/tolerances stated": ["error bound", "tolerance", "accuracy", "residual"],
        "Robustness under worst-case/edge conditions": ["worst case", "worst-case", "stress", "edge case",
                                                        "corner case", "boundary condition"],
        "Deterministic behavior across platforms": ["deterministic", "platform", "portability", "repeatable"],
        "Exception/NaN/Inf handling defined": ["nan", "inf", "exception"],
        "Algorithm stability/conditioning addressed": ["stability", "numerical stability", "conditioning",
                                                       "ill-conditioned", "converge", "diverge"],
        "Test coverage for numeric boundaries": ["test", "unit test", "verification", "validation", "boundary"]
    }

    check_rx = {
        checks: [re.compile(r'\b' + re.escape(term) + r'\b', re.IGNORECASE) for term in terms]
        for checks, terms in checks.items()
    }

    for index, row in algo_df.iterrows():
        desc_text = str(row.get('DESC', '') or '')
        req_id = row.get('ID', f"Row{index}")

        if not desc_text.strip():
            results['details'].append(f"Notice: Empty DESC at ID [{req_id}]")
            continue

        txt = ' ' + desc_text.lower() + ' '
        matched = {}
        for checks, patterns in check_rx.items():
            terms_hit = []
            for rx in patterns:
                if rx.search(txt):
                    term_literal = rx.pattern.replace(r'\b', '')
                    terms_hit.append(term_literal)
            if terms_hit:
                matched[checks] = sorted(set(terms_hit))

        if matched:
            if results['status'] == "Passed":
                results['status'] = "Failed"
            checks_list = ';'.join(matched.keys())
            term_detail = '; '.join([f"{k}: {', '.join(v)}" for k, v in matched.items()])
            results['details'].append(
                f"Notice: Edge-case categories hit at ID [{req_id}] → [{checks_list}] | terms: {term_detail}"
            )

    if results['status'] == "Passed":
        results['details'].append(
            f"Scanned {total_algos} requirements; no DO-178 edge-case terms found."
        )
    append_result(results)


## Objective G 7.7 ##

def div_by_zero_check(algo_df):
    """
    Detect literal division-by-zero patterns.
    """
    results = {
        'check': "div_by_zero_check",
        'status': "Passed",
        'details': [],
        'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'ids': list(pd.unique(algo_df['ID']))  # <-- Per-ID tracking
    }

    total_algos = len(algo_df)
    if total_algos == 0:
        results['status'] = "Failed: algo_df is empty"
        append_result(results)
        ##return results

    div_zero_pattern = r'/\s*0(?:\.0+)?(?!\d)'

    for index, row in algo_df.iterrows():
        desc_text = str(row.get('DESC', ''))
        req_id = row.get('ID', index)
        lines = desc_text.splitlines()
        for line_num, line_content in enumerate(lines, 1):
            if '/' in line_content and re.search(div_zero_pattern, line_content):
                error_msg = f"Error: Zero Division Hazard at ID [{req_id}], Line {line_num}: '{line_content.strip()}'"
                results['status'] = "Failed"
                results['details'].append(error_msg)

    if not results['details']:
        results['details'].append(f"Scanned {total_algos} requirements; no literal division by zero found.")
    append_result(results)
    ##return results


def infinite_loop_check(algo_df):
    """
    Detect common infinite-loop pseudocode markers (while(1), while True, typo while Ture).
    """
    results = {
        'check': "infinite_loop_check",
        'status': "Passed",
        'details': [],
        'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'ids': list(pd.unique(algo_df['ID']))  # <-- Per-ID tracking
    }

    total_algos = len(algo_df)
    if total_algos == 0:
        results['status'] = "Failed: algo_df is empty"
        append_result(results)
        ##return results

    combined_pattern = r'(while\s*\(\s*1\s*\)|while\s*True|while\s*Ture)'

    for index, row in algo_df.iterrows():
        desc_text = str(row.get('DESC', ''))
        req_id = row.get('ID', index)
        lines = desc_text.splitlines()
        for line_num, line_content in enumerate(lines, 1):
            if re.search(combined_pattern, line_content, re.IGNORECASE):
                error_msg = f"Caution: Infinite Loop Hazard at ID [{req_id}], Line {line_num}: '{line_content.strip()}'"
                results['status'] = "Failed"
                results['details'].append(error_msg)

    if not results['details']:
        results['details'].append(f"Scanned {total_algos} requirements; no infinite loop patterns detected.")
    append_result(results)
    ##return results


def null_pointer_check(algo_df):
    """
    Detect null pointer textual mentions and actual None/NaN cells.
    """
    results = {
        'check': "null_pointer_check",
        'status': "Passed",
        'details': [],
        'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'ids': list(pd.unique(algo_df['ID']))  # <-- Per-ID tracking
    }

    total_algos = len(algo_df)
    if total_algos == 0:
        results['status'] = "Failed: algo_df is empty"
        append_result(results)
        ##return results

    patterns = [
        r"\bnull\s*pointer\b",
        r"\bnullptr\b",
        r"\bnull\s*ptr\b",
        r"\b\w+\s*=\s*null\b",
        r"\b\w+\s*<-\s*null\b",
        r"\bdeclare\s+\w+\s*=\s*null\b"
    ]

    for index, row in algo_df.iterrows():
        req_id = row.get('ID', f"Row-{index}")

        for col in algo_df.columns:
            if pd.isnull(row[col]):
                error_msg = f"Error: Null value detected at ID [{req_id}], Column '{col}'"
                results['status'] = "Failed"
                results['details'].append(error_msg)

        desc_text = str(row.get('DESC', ""))
        lines = desc_text.splitlines()
        for line_num, line_content in enumerate(lines, 1):
            for pattern in patterns:
                if re.search(pattern, line_content, re.IGNORECASE):
                    error_msg = f"Caution: Null Pointer mention at ID [{req_id}], Line {line_num}: '{line_content.strip()}'"
                    results['status'] = "Failed"
                    if error_msg not in results['details']:
                        results['details'].append(error_msg)
                    break

    if not results['details']:
        results['details'].append(f"Scanned {total_algos} requirements; no null pointers detected.")
    append_result(results)
    ##return results


def out_of_range_check(algo_df):
    """
    Detect out-of-range array or list access in pseudocode stored in algo_df.
    """
    results = {
        'check': "out_of_range_check",
        'status': "Passed",
        'details': [],
        'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'ids': list(pd.unique(algo_df['ID']))  # <-- Per-ID tracking
    }

    total_algos = len(algo_df)
    if total_algos == 0:
        results['status'] = "Failed: algo_df is empty"
        append_result(results)
        ##return results

    # Cleaned regex patterns for suspicious array access
    patterns = [
        r"\barray\s*\[\s*\d+\s*\]",  # array[10]
        r"\barray\s*\[\s*index\s*\]",
        r"\barray\s*\[\s*i\s*\]",
        r"\barray\s*\[\s*j\s*\]",
        r"\barray\s*\[\s*k\s*\]"
    ]

    array_decl_pattern = r"SET\s+array\s*=\s*\[(.* ?)\]"
    loop_pattern = r"FOR\s+(\w+)\s*=\s*(\d+)\s*to\s*(\d+)"

    for index, row in algo_df.iterrows():
        req_id = row.get('ID', f"Row-{index}")
        desc_text = str(row.get('DESC', ""))
        lines = desc_text.splitlines()

        array_length = None
        decl_match = re.search(array_decl_pattern, desc_text, re.IGNORECASE)
        if decl_match:
            elements = decl_match.group(1).split(",")
            array_length = len([e.strip() for e in elements if e.strip()])

        for line_num, line_content in enumerate(lines, 1):
            for pattern in patterns:
                if re.search(pattern, line_content, re.IGNORECASE):
                    error_msg = f"Notice: Potential array access at ID [{req_id}], Line {line_num}: '{line_content.strip()}'"
                    if error_msg not in results['details']:
                        results['status'] = "Failed"
                        results['details'].append(error_msg)
                    break

            loop_match = re.search(loop_pattern, line_content, re.IGNORECASE)
            if loop_match and array_length is not None:
                start_idx = int(loop_match.group(2))
                end_idx = int(loop_match.group(3))
                if end_idx >= array_length:
                    error_msg = (
                        f"Error: Out-of-range loop detected at ID [{req_id}], Line {line_num}: "
                        f"Loop goes to {end_idx} but array length is {array_length}"
                    )
                    if error_msg not in results['details']:
                        results['status'] = "Failed"
                        results['details'].append(error_msg)

    if not results['details']:
        results['details'].append(
            f"Scanned {total_algos} requirements; no input abnormalities patterns detected."
        )
    append_result(results)
    ##return results


def var_input_analysis(algo_df):
    """
    Placeholder for input analysis; currently no checks.
    """
    results = {
        'check': "var_input_analysis",
        'status': "Passed",
        'details': [],
        'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'ids': list(pd.unique(algo_df['ID']))  # <-- Per-ID tracking (so IDs show as Passed)
    }

    total_algos = len(algo_df)
    if total_algos == 0:
        results['status'] = "Failed: algo_df is empty"
        append_result(results)
        ##return results

    results['details'].append(f"Scanned {total_algos} records; no input abnormalities patterns detected.")
    append_result(results)
    ##return results
