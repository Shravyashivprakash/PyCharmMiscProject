import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from config import G4_BASE_INPUT, BASE_OUTPUT, G4_OUTPUT_FILE
from io_utils import g4_collect_requirements
import g4.g4_logic as g4_logic


def run_g4():
    os.makedirs(BASE_OUTPUT, exist_ok=True)
    output_file = os.path.join(BASE_OUTPUT, G4_OUTPUT_FILE)

    # ================= READ INPUTS =================
    all_requirements = g4_collect_requirements(G4_BASE_INPUT)

    # ================= EXCEL OUTPUT =================
    wb = Workbook()
    ws = wb.active
    ws.title = "HLR's are verifiable"

    ws.merge_cells("A1:J1")
    ws["A1"] = "High-Level Requirements are Verifiable"
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        "Req ID",
        "G 4.1/4.2 Check testability: Confirm each requirement can be verified through testing or analysis.",
        "G 4.3 Verification Method (Declared)",
        "G 4.5 Use mandatory language: Confirm 'shall' is used for mandatory requirements.",
        "Register Requirement",
        "Communication Requirement",
        "Fault Requirement",
        "I/O Requirement",
        "Functional Requirement"
    ]

    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="bottom")

    ws.freeze_panes = "A3"

    # ================= DATA =================
    for rid, rtxt in all_requirements:
        ws.append([
            rid,
            g4_logic.check_g41_testability(rtxt),
            g4_logic.extract_verification_method(rtxt),
            g4_logic.check_g45(rtxt),
            "Yes" if g4_logic.is_register_requirement(rtxt) else "No",
            "Yes" if g4_logic.is_communication_requirement(rtxt) else "No",
            "Yes" if g4_logic.is_fault_requirement(rtxt) else "No",
            "Yes" if g4_logic.is_io_requirement(rtxt) else "No",
            "Yes" if g4_logic.is_functional_requirement(rtxt) else "No",
        ])

    # ================= FORMATTING =================
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")

    for row in ws.iter_rows(min_row=3, max_col=len(headers)):
        for cell in row:
            if cell.value in ("PASS", "Yes"):
                cell.fill = green
            elif cell.value in ("FAIL", "No"):
                cell.fill = red

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)

    wb.save(output_file)

    print("\n✅ G4 review completed")
    print("📄 Output:", output_file)
