from pathlib import Path
from collections import defaultdict
from datetime import datetime

from io_utils import extract_docx_text, extract_requirements
from g5.g5_logic import run_g5_checks

from config import (
    BASE_INPUT,
    BASE_OUTPUT,
    INPUT_SRS_FILE,
    G5_OUTPUT_FILE,
    REPORT_TITLE,
    COLUMN_HEADERS
)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def generate_excel_report(all_req_ids, findings_by_req, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "SRS Review"

    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(COLUMN_HEADERS))
    ws["A1"] = REPORT_TITLE
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append(COLUMN_HEADERS)
    for col in range(1, len(COLUMN_HEADERS) + 1):
        ws.cell(row=2, column=col).font = Font(bold=True)
        ws.cell(row=2, column=col).alignment = Alignment(wrap_text=True)

    row = 3
    for req_id in sorted(all_req_ids):
        issues = findings_by_req.get(req_id, [])

        status = {k: "PASS" for k in ["5.1", "5.2", "5.3", "5.4", "5.5"]}
        for issue in issues:
            for k in status:
                if k in issue:
                    status[k] = "FAIL"

        ws.append([
            req_id,
            status["5.1"],
            status["5.2"],
            status["5.3"],
            status["5.4"],
            status["5.5"],
            "\n".join(dict.fromkeys(issues)) if issues else "N/A"
        ])

        ws.cell(row=row, column=7).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        row += 1

    try:
        wb.save(output_path)
        print(f"\n✅ G5 Excel generated: {output_path}")
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = output_path.with_stem(output_path.stem + "_" + ts)
        wb.save(fallback)
        print(f"\n⚠ File open. Saved as:\n{fallback}")


def run_g5():
    srs_path = Path(BASE_INPUT) / INPUT_SRS_FILE
    output_path = Path(BASE_OUTPUT) / G5_OUTPUT_FILE

    # ✅ STEP 1: read DOCX → TEXT
    srs_text = extract_docx_text(srs_path)

    # ✅ STEP 2: extract requirements (G1-style)
    raw_reqs = extract_requirements(srs_text)

    # ✅ STEP 3: adapt to G5 expected structure
    requirements = [
        {
            "id": rid,
            "raw_id": rid,
            "text": text
        }
        for rid, text in raw_reqs.items()
    ]

    # ✅ STEP 4: run G5 logic
    g5_results, all_findings = run_g5_checks(requirements)

    # ✅ STEP 5: build master ID list
    valid_ids = {r["id"] for r in requirements if r.get("id")}
    reported_ids = {req_id for req_id, _ in all_findings if req_id}
    all_req_ids = valid_ids | reported_ids

    # ✅ STEP 6: group findings
    findings_by_req = defaultdict(list)
    for req_id, issue in all_findings:
        findings_by_req[req_id].append(issue)

    # ✅ STEP 7: generate Excel
    generate_excel_report(all_req_ids, findings_by_req, output_path)

    print("<<< G5 review completed")

